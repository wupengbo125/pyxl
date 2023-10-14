import openpyxl
from openpyxl.styles import PatternFill

# 打开Excel文件
workbook = openpyxl.load_workbook('fruits.xlsx')

# 选择工作表
worksheet = workbook.active

# 创建填充背景色的样式
odd_fill = PatternFill(start_color='D9D9D9', end_color='FF9999', fill_type='solid')
even_fill = PatternFill(start_color='D9D9D9', end_color='A4C2B7', fill_type='solid')


# 从第五行第四列开始遍历单元格
for row in worksheet.iter_rows(min_row=5, min_col=4):
    for col_index, cell in enumerate(row, start=4):
        # 检查单元格中的文本是否大于10个字符
        if len(str(cell.value)) > 10:
            if col_index % 2 == 0:
                cell.fill = even_fill
            else:
                cell.fill = odd_fill

# 保存修改后的Excel文件
workbook.save('fruits.xlsx')

# 关闭工作簿
workbook.close()
