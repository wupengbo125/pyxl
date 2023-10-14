import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from css import css

# 加载工作簿和工作表
wb = load_workbook('fruits.xlsx')
ws = wb.active

# 参数列表
params = sys.argv[1:]  # 获取命令行参数

# 从第4行第4列开始搜索
start_row = 4
start_col = 4

# 创建新的工作表并将数据复制过去
new_ws = wb.create_sheet(title=" ".join(params))
for row in ws.iter_rows(values_only=True):
    new_ws.append(row)

# 需要删除的列
cols_to_delete = []

# 遍历每一列
for col in range(start_col, new_ws.max_column + 1):
    # 假设该列需要被删除
    delete_col = True

    # 在第4行的该列中搜索参数
    cell_value = new_ws.cell(row=start_row, column=col).value
    if cell_value is None:
        print("存在标签为空的列，请补充所有经方的标签")
        break
    if all(param in cell_value for param in params):
        # 如果找到所有参数，那么该列不应被删除
        delete_col = False

    # 如果该列需要被删除，将其添加到删除列表中
    if delete_col:
        cols_to_delete.append(col)

# 按照从大到小的顺序删除列，这样就不会影响其他列的索引
for col in sorted(cols_to_delete, reverse=True):
    new_ws.delete_cols(col)

# 调用设置样式的函数
css(new_ws)

# 保存工作簿为'fruits.xlsx'
wb.save('fruits.xlsx')
