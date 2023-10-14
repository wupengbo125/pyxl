from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def css(ws):
    # 设置新工作表的格式为垂直靠上，wrap为真，微软雅黑八号
    font = Font(name='微软雅黑', size=8)
    alignment = Alignment(vertical='top', wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    # 设置第一行奇数列为红色，偶数列为绿色
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="a4c2b7", end_color="a4c2b7", fill_type="solid")

    for i, cell in enumerate(ws[1], start=1):
        if i % 2 == 0:  # 偶数列
            cell.fill = green_fill
        else:  # 奇数列
            cell.fill = red_fill

    # 设置第二行和第三行为微软雅黑7号，橙色
    font = Font(name='微软雅黑', size=7, color="FFA500")
    for row in ws.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.font = font

    # 设置所有行宽为10，高为13
    for i in range(1,ws.max_column +1):
        ws.column_dimensions[get_column_letter(i)].width = 10
    for i in range(1,ws.max_row + 1):
        ws.row_dimensions[i].height = 13