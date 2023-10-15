from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from css import css
import argparse

# 放心大胆的用，该方法很强大，能处理单行药方，双行说明加药方，三行说明加药方加说明，五行三行加替代方加药方。5n行，
count_start_line = 5
count_start_column = 4
jfnamefrom = -6 # 数据源的最后几个字是经方名字吗？从哪开始
jfnameend = -1 # 从哪结束
jfname_line = 1 # 经方名放在第几行
question_line = 2 # 人有
jfexplain_line = 3 # 盖
label_line = 4 # label

# 创建ArgumentParser对象
parser = argparse.ArgumentParser(description="Add a parameter to the fourth row of each new column.")
parser.add_argument("parameter", nargs='?', default='', help="The parameter to add to the fourth row of each new column.")

# 解析命令行参数
args = parser.parse_args()

# 定义一个函数来处理特定的行
def process_line(ws, fruits, last_column, font, alignment):
    # 遍历每个水果
    for fruit in fruits:
        # 检查水果名称的前两个字符是否在Excel的第一列中
        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            if fruit[:2] in row:
                # 如果找到了，就写在那一行的最后一列，并设置字体和对齐方式
                cell = ws.cell(row=i, column=last_column, value=fruit)
                cell.font = font
                cell.alignment = alignment
                break
        else:
            # 如果没找到，就把那个水果的前两个字写到第一列最后一个位置，再把他写到最后一列的那个位置，并设置字体和对齐方式
            ws.append({1: fruit[:2], last_column: fruit})
            new_row = ws[ws.max_row]
            for cell in new_row:
                cell.font = font
                cell.alignment = alignment

# 读取Excel文件
wb = load_workbook('fruits.xlsx')
ws = wb.active

# 创建一个Font对象，设置字体为微软雅黑，字号为8
font = Font(name='微软雅黑', size=8)

# 创建一个Alignment对象，设置文本居中显示和自动换行
alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

# 读取文本文件，使用utf-8编码
with open('bzl.txt', 'r', encoding='utf-8') as file:
    lines = file.readlines()

# 过滤掉空行
lines = [line for line in lines if line.strip() != '']

# 获取Excel文件的最后一列
last_column = ws.max_column + 1

# 如果文件只有一行，将这一行视为三行模式下的第二行进行处理
if len(lines) == 1:
    # 对这一行进行split，并调用process_line函数处理
    fruits = lines[0].split('，')
    process_line(ws, fruits, last_column, font, alignment)
    # 在第四行添加参数的值
    if args.parameter:
        ws.cell(row=label_line, column=last_column, value=args.parameter)
elif len(lines) == 2:
    # 处理the 1st line
    cell1 = ws.cell(row=jfname_line, column=last_column, value=lines[0].strip()[jfnamefrom:jfnameend])
    cell1 = ws.cell(row=question_line, column=last_column, value=lines[0].strip())

    # 对第2行进行split，并调用process_line函数处理
    fruits = lines[1].split('，')
    process_line(ws, fruits, last_column, font, alignment)
    # 在第四行添加参数的值
    if args.parameter:
        ws.cell(row=label_line, column=last_column, value=args.parameter)
# 如果文件有三行或者五行或者其他5的倍数行
elif len(lines) >= 3:
    # 将文件拆分成多个五行进行处理
    for i in range(0, len(lines), 5):
        chunk = lines[i:i+5]
        # 处理前三行或者只有一行的情况
        cell0 = ws.cell(row=jfname_line, column=last_column, value=chunk[0].strip()[jfnamefrom:jfnameend])
        cell1 = ws.cell(row=question_line, column=last_column, value=chunk[0].strip())

        if len(chunk) > 1:
            cell2 = ws.cell(row=3, column=last_column, value=chunk[2].strip())
            # 对第二行进行split，并调用process_line函数处理
            fruits = chunk[1].split('，')
            process_line(ws, fruits, last_column, font, alignment)
            # 在第四行添加参数的值
            if args.parameter:
                ws.cell(row=label_line, column=last_column, value=args.parameter)

        if len(chunk) > 3:
            # 处理第四行和第五行
            last_column += 1
            celln3 = ws.cell(row=jfname_line, column=last_column, value=chunk[3].strip()[jfnamefrom:jfnameend])
            cell3 = ws.cell(row=question_line, column=last_column, value=chunk[3].strip())

            # 在第四行添加参数的值
            if args.parameter:
                ws.cell(row=label_line, column=last_column, value=args.parameter)

            # 对第五行进行split，并调用process_line函数处理
            fruits = chunk[4].split('，')
            process_line(ws, fruits, last_column, font, alignment)

        last_column += 1
# count
for i in range(count_start_line, ws.max_row + 1):
    count = sum(1 for cell in ws[i][count_start_column-1:] if cell.value is not None)
    ws.cell(row=i, column=2, value=count)

css(ws)

# 保存更改后的Excel文件
wb.save('fruits.xlsx')