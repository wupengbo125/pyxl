import openpyxl
from openpyxl.styles import Alignment, Font

# 打开工作簿和工作表
wb = openpyxl.load_workbook('fruits.xlsx')
sheet = wb.active

# 用utf8编码读取文本文件
with open('xwgj.txt', 'r', encoding='utf8') as f:
    text = f.read()

# 在Excel的第一列中搜索文本文件的前两个字符
for row in range(1, sheet.max_row + 1):
    if sheet.cell(row=row, column=1).value == text[:2]:
        # 如果找到匹配项，将整个文本内容写入到第三列的相应行
        cell = sheet.cell(row=row, column=3)
        cell.value = text
        # 设置wraptext为true
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        # 设置字体为微软雅黑，字号为8
        cell.font = Font(name='微软雅黑', size=8)

# 保存更改
wb.save('fruits.xlsx')
